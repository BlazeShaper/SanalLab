"""
/api/reports/* — Report builder endpoints.
"""
from __future__ import annotations
import datetime
import uuid

from fastapi import APIRouter, Depends, Request, Response
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db import get_db
from app.models import ReportItem
from app.session import ensure_session

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---- schemas ----------------------------------------------------------

class AddItemPayload(BaseModel):
    expression: str = "coulomb"
    label: str = "Coulomb's Law: F = k · (|q₁|·|q₂|) / r²"
    q1: float = 0.0
    q2: float = 0.0
    r: float = 1.0
    force: float = 0.0


class ReorderPayload(BaseModel):
    fromIndex: int
    toIndex: int


# ---- endpoints --------------------------------------------------------

@router.get("/items")
def list_items(request: Request, response: Response, db: Session = Depends(get_db)):
    sid = ensure_session(request, response)
    items = (
        db.query(ReportItem)
        .filter(ReportItem.session_id == sid)
        .order_by(ReportItem.order_index)
        .all()
    )
    return [_ser(it) for it in items]


@router.post("/items/add")
def add_item(payload: AddItemPayload, request: Request, response: Response, db: Session = Depends(get_db)):
    sid = ensure_session(request, response)

    count = db.query(ReportItem).filter(ReportItem.session_id == sid).count()

    item = ReportItem(
        id=str(uuid.uuid4()),
        session_id=sid,
        expression=payload.expression,
        label=payload.label,
        q1=payload.q1,
        q2=payload.q2,
        r=payload.r,
        force=payload.force,
        order_index=count,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _ser(item)


@router.post("/items/reorder")
def reorder_items(payload: ReorderPayload, request: Request, response: Response, db: Session = Depends(get_db)):
    sid = ensure_session(request, response)
    items = (
        db.query(ReportItem)
        .filter(ReportItem.session_id == sid)
        .order_by(ReportItem.order_index)
        .all()
    )
    if 0 <= payload.fromIndex < len(items) and 0 <= payload.toIndex < len(items):
        moved = items.pop(payload.fromIndex)
        items.insert(payload.toIndex, moved)
        for idx, it in enumerate(items):
            it.order_index = idx
        db.commit()
    return [_ser(it) for it in items]


@router.delete("/items/{item_id}")
def delete_item(item_id: str, request: Request, response: Response, db: Session = Depends(get_db)):
    sid = ensure_session(request, response)
    item = db.query(ReportItem).filter(ReportItem.id == item_id, ReportItem.session_id == sid).first()
    if item:
        db.delete(item)
        db.commit()
    return {"ok": True}


@router.get("/export/html", response_class=HTMLResponse)
def export_html(
    request: Request,
    response: Response,
    title: str = "İnteraktif Fizik Lab Raporu",
    lang: str = "tr",
    db: Session = Depends(get_db),
):
    sid = ensure_session(request, response)
    items = (
        db.query(ReportItem)
        .filter(ReportItem.session_id == sid)
        .order_by(ReportItem.order_index)
        .all()
    )

    rows = ""
    for idx, it in enumerate(items):
        rows += (
            f"<tr><td>{idx + 1}</td><td>{_esc(it.label)}</td>"
            f"<td>{it.q1:.1f} µC</td><td>{it.q2:.1f} µC</td>"
            f"<td>{it.r:.2f} m</td><td>{it.force:.3f} N</td>"
            f"<td>{it.created_at}</td></tr>\n"
        )

    # i18n labels for export
    _labels = {
        "tr": {
            "created_at": "Oluşturulma tarihi: ",
            "saved_expressions": "Kaydedilmiş İfadeler",
            "expression": "İfade",
            "timestamp": "Zaman Damgası",
            "no_items": "Kaydedilmiş öğe yok.",
            "hint": 'PDF olarak dışa aktar: Tarayıcı Yazdır → "PDF olarak Kaydet".',
        },
        "en": {
            "created_at": "Created at: ",
            "saved_expressions": "Saved Expressions",
            "expression": "Expression",
            "timestamp": "Timestamp",
            "no_items": "No items saved.",
            "hint": 'Export as PDF: Browser Print → "Save as PDF".',
        },
    }
    lb = _labels.get(lang, _labels["tr"])

    if not rows:
        rows = f'<tr><td colspan="7">{lb["no_items"]}</td></tr>'

    created = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html = f"""<!doctype html>
<html lang="{lang}"><head><meta charset="utf-8"/><title>{_esc(title)}</title>
<style>
body{{font-family:Arial,sans-serif;margin:24px;color:#111}}
h1{{margin:0 0 8px}}
.meta{{color:#555;margin-bottom:16px;font-size:13px}}
table{{width:100%;border-collapse:collapse;margin-top:12px}}
th,td{{border:1px solid #ddd;padding:8px;font-size:13px;vertical-align:top}}
th{{background:#f5f5f5;text-align:left}}
.hint{{margin-top:16px;color:#666;font-size:12px}}
</style></head><body>
<h1>{_esc(title)}</h1>
<div class="meta">{lb['created_at']}{created}</div>
<h2>{lb['saved_expressions']}</h2>
<table><thead><tr><th>#</th><th>{lb['expression']}</th><th>q1</th><th>q2</th><th>r</th><th>F</th><th>{lb['timestamp']}</th></tr></thead>
<tbody>{rows}</tbody></table>
<div class="hint">{lb['hint']}</div>
</body></html>"""

    import re
    ascii_title = re.sub(r'[^a-zA-Z0-9_\- ]', '', title).strip().replace(" ", "_")
    if not ascii_title:
        ascii_title = "report"
    
    return HTMLResponse(content=html, headers={
        "Content-Disposition": f'attachment; filename="{ascii_title}.html"'
    })


@router.get("/export/json")
def export_json(request: Request, response: Response, db: Session = Depends(get_db)):
    sid = ensure_session(request, response)
    items = (
        db.query(ReportItem)
        .filter(ReportItem.session_id == sid)
        .order_by(ReportItem.order_index)
        .all()
    )
    return JSONResponse(content=[_ser(it) for it in items])


@router.get("/export/xlsx")
def export_xlsx(
    request: Request,
    response: Response,
    title: str = "İnteraktif Fizik Lab Raporu",
    lang: str = "tr",
    db: Session = Depends(get_db),
):
    """Generate and return an Excel (.xlsx) report."""
    import io
    from urllib.parse import quote
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sid = ensure_session(request, response)
    items = (
        db.query(ReportItem)
        .filter(ReportItem.session_id == sid)
        .order_by(ReportItem.order_index)
        .all()
    )

    # i18n headers
    _hdr = {
        "tr": {
            "num": "#",
            "expression": "İfade",
            "q1": "q1 (µC)",
            "q2": "q2 (µC)",
            "r": "r (m)",
            "force": "F (N)",
            "timestamp": "Zaman Damgası",
            "sheet": "Rapor",
            "created": "Oluşturulma",
        },
        "en": {
            "num": "#",
            "expression": "Expression",
            "q1": "q1 (µC)",
            "q2": "q2 (µC)",
            "r": "r (m)",
            "force": "F (N)",
            "timestamp": "Timestamp",
            "sheet": "Report",
            "created": "Created",
        },
    }
    h = _hdr.get(lang, _hdr["tr"])

    wb = Workbook()
    ws = wb.active
    ws.title = h["sheet"]

    # ── Title row ──
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="EC5B13")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Created-at row ──
    ws.merge_cells("A2:G2")
    meta_cell = ws["A2"]
    meta_cell.value = f"{h['created']}: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    meta_cell.font = Font(size=10, italic=True, color="666666")
    meta_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── Header row (row 4) ──
    headers = [h["num"], h["expression"], h["q1"], h["q2"], h["r"], h["force"], h["timestamp"]]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="EC5B13", end_color="EC5B13", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for col_idx, hdr_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=hdr_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 22

    # ── Data rows ──
    data_font = Font(size=11)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left = Alignment(horizontal="left", vertical="center")
    alt_fill = PatternFill(start_color="FFF3ED", end_color="FFF3ED", fill_type="solid")

    for idx, it in enumerate(items):
        row_num = 5 + idx
        row_data = [
            idx + 1,
            it.label,
            round(it.q1, 2),
            round(it.q2, 2),
            round(it.r, 4),
            round(it.force, 6),
            str(it.created_at),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align_center if col_idx != 2 else data_align_left
            if idx % 2 == 1:
                cell.fill = alt_fill

    # ── Column widths ──
    col_widths = [5, 40, 12, 12, 12, 16, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── Freeze header ──
    ws.freeze_panes = "A5"

    # ── Write to buffer ──
    buf = io.BytesIO()
    wb.save(buf)

    # Convert to pure ascii to prevent enterprise browser/proxy rejection
    import re
    ascii_title = re.sub(r'[^a-zA-Z0-9_\- ]', '', title).strip().replace(" ", "_")
    if not ascii_title:
        ascii_title = "report"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_title}.xlsx"'},
    )


# ---- helpers ----------------------------------------------------------

def _ser(item: ReportItem) -> dict:
    return {
        "id": item.id,
        "expression": item.expression,
        "label": item.label,
        "q1": item.q1,
        "q2": item.q2,
        "r": item.r,
        "force": item.force,
        "order_index": item.order_index,
        "created_at": str(item.created_at),
    }


def _esc(s: str) -> str:
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
